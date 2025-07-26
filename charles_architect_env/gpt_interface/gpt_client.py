import os
import fitz  # PyMuPDF
import openpyxl
from openai import OpenAI
from openai import OpenAIError
from dotenv import load_dotenv

# טען משתני סביבה מהקובץ .env
load_dotenv()

# יצירת לקוח OpenAI עם המפתח מה־.env
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))


def load_context_from_documents(folder_path="project_docs", max_chars=12000):
    """
    טוען תוכן מכל המסמכים הנתמכים בתיקייה (PDF, Excel, TXT)
    """
    context_text = ""

    for filename in os.listdir(folder_path):
        filepath = os.path.join(folder_path, filename)
        ext = filename.lower().split(".")[-1]

        try:
            if ext == "txt":
                with open(filepath, "r", encoding="utf-8") as f:
                    content = f.read()
                    context_text += f"\n\n📄 {filename}:\n" + content

            elif ext == "pdf":
                with fitz.open(filepath) as doc:
                    text = "\n".join(page.get_text() for page in doc)
                    context_text += f"\n\n📄 {filename} (PDF):\n" + text

            elif ext in ["xlsx", "xls"]:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                for sheet in wb.sheetnames:
                    sheet_obj = wb[sheet]
                    rows = []
                    for row in sheet_obj.iter_rows(values_only=True):
                        line = "\t".join([str(cell) if cell is not None else "" for cell in row])
                        rows.append(line)
                    context_text += f"\n\n📊 {filename} – גליון {sheet}:\n" + "\n".join(rows[:20])  # רק 20 שורות ראשונות

        except Exception as e:
            context_text += f"\n⚠️ שגיאה בקריאת {filename}: {str(e)}\n"

    return context_text[:max_chars]


def ask_gpt(prompt: str):
    """
    שואל את GPT שאלה עם קונטקסט מותאם אישי מהמסמכים
    """
    system_prompt = (
        "אתה Charles_Architect – GPT מומחה לפיתוח מערכות AI מתקדמות לזיהוי מניות פורצות. "
        "המטרה שלך היא להכיר את המערכת ש־רן בונה, ללמוד את כל המסמכים והקבצים, ולייעץ, לבנות קוד, להציע תשתית, ולחשוב קדימה. "
        "ענה בעברית ברורה, שלב־אחרי־שלב, כאילו אתה שותף בצוות. כל שאלה תתייחס תמיד לקונטקסט שנטען מהמסמכים של המערכת."
    )

    context = load_context_from_documents()
    full_prompt = f"{context}\n\n🔍 שאלה:\n{prompt}"

    try:
        response = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": full_prompt}
            ],
            temperature=0.4,
            max_tokens=1000
        )
        return response.choices[0].message.content.strip()

    except OpenAIError as e:
        return f"שגיאה בבקשת GPT: {str(e)}"
